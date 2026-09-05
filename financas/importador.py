"""
importador.py — O caminho completo de um arquivo ate o banco.
==============================================================================

O QUE ACONTECE QUANDO VOCE IMPORTA UM ARQUIVO

    1. DETECTAR   que tipo de arquivo e este? (fatura, extrato CSV, OFX)
    2. LER        transformar o arquivo em linhas normalizadas (leitores/)
    3. IDENTIFICAR criar a impressao digital de cada linha
    4. DEDUPLICAR  marcar o que ja existe no banco como DUPLICADO
    5. CLASSIFICAR aplicar as regras para sugerir categoria (regras.py)
    6. REVISAR     voce confere e corrige na tela  <- a unica etapa manual
    7. GRAVAR      so o que esta marcado como NOVO entra no banco

A ETAPA 4 E A QUE MAIS IMPORTA
------------------------------
Seus extratos SE SOBREPOEM. O CSV vai de 08/05 a 06/08 e o OFX vai de 22/07 a
21/08 — os dias entre 22/07 e 06/08 estao NOS DOIS ARQUIVOS. Sem deduplicacao,
importar os dois faria duas semanas de gastos contarem em dobro, e todo numero
do painel ficaria errado sem nenhum aviso.

COMO CADA TIPO E IDENTIFICADO

    OFX          pelo FITID — um codigo unico que o proprio banco atribui.
                 E o melhor caso: nao ha adivinhacao nenhuma.

    Extrato CSV  data + hora + descricao + valor. A HORA e o que salva: dois
                 Pix de R$ ···· para a mesma pessoa no mesmo dia sao
                 diferentes se sairam em minutos diferentes.

    Fatura CSV   mes + data + estabelecimento + portador + valor + parcela,
                 MAIS um contador de ocorrencia. O contador existe porque
                 duas corridas de Uber de R$ ···· no mesmo dia sao duas
                 corridas de verdade — sem ele, a segunda sumiria.

DUAS CAMADAS DE PROTECAO
------------------------
Alem da checagem aqui, a coluna id_unico da tabela e UNIQUE e o INSERT usa
"OR IGNORE". Ou seja: mesmo que este codigo tenha uma falha, o proprio banco
recusa a duplicata. Nunca confie numa camada so para proteger seus dados.
"""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd

from financas import banco, config, regras
from financas.formato import chave_hash, hash_arquivo
from financas.leitores import (extrato_csv, extrato_ofx, extrato_xp_xlsx,
                               fatura_csv, posicao_xp)
from financas.leitores.base import ResultadoLeitura

STATUS_NOVO = "NOVO"
STATUS_DUPLICADO = "DUPLICADO"

TIPO_FATURA = "fatura"
TIPO_EXTRATO_CSV = "extrato_csv"
TIPO_OFX = "ofx"
TIPO_POSICAO_XP = "posicao_xp"
TIPO_EXTRATO_XP = "extrato_xp"


def detectar_tipo(nome_arquivo: str, amostra: str = "") -> str | None:
    """Descobre que tipo de arquivo e, pelo nome e pelo conteudo.

    Olha o CONTEUDO tambem, e nao so a extensao, porque arquivo baixado do
    banco costuma vir com nome trocado ("extrato (1).csv", "download.ofx").
    O cabecalho nunca mente.
    """
    nome = (nome_arquivo or "").lower()
    cabecalho = (amostra or "")[:1000].lower()

    if nome.endswith(".ofx") or "<ofx" in cabecalho or "ofxheader" in cabecalho:
        return TIPO_OFX
    if "estabelecimento" in cabecalho or "portador" in cabecalho:
        return TIPO_FATURA
    if "saldo" in cabecalho and "descricao" in cabecalho:
        return TIPO_EXTRATO_CSV
    if nome.startswith("fatura"):
        return TIPO_FATURA
    if nome.startswith("extrato"):
        return TIPO_EXTRATO_CSV
    return None


def ler(caminho=None, dados_bytes: bytes | None = None, nome_arquivo: str = "",
        mes_competencia: str | None = None) -> tuple[ResultadoLeitura, str | None]:
    """Le um arquivo de qualquer um dos tres tipos.

    Aceita as duas formas de entrada:
        - `caminho`: um arquivo em disco (usado pela importacao por pasta)
        - `dados_bytes` + `nome_arquivo`: o upload do Streamlit, que entrega
          o arquivo como bytes na memoria

    Devolve (resultado, tipo_detectado).
    """
    if caminho is not None:
        caminho = Path(caminho)
        nome_arquivo = nome_arquivo or caminho.name
        dados_bytes = caminho.read_bytes()

    if dados_bytes is None:
        resultado = ResultadoLeitura()
        resultado.erros.append("Nenhum arquivo foi fornecido.")
        return resultado, None

    if dados_bytes[:2] == b"PK" or (nome_arquivo or "").lower().endswith(".xlsx"):
        tipo = detectar_tipo_excel(dados_bytes, nome_arquivo)
        if tipo == TIPO_POSICAO_XP:
            return posicao_xp.ler_bytes(dados_bytes, nome_arquivo), tipo
        if tipo == TIPO_EXTRATO_XP:
            return extrato_xp_xlsx.ler_bytes(dados_bytes, nome_arquivo), tipo
        resultado = ResultadoLeitura()
        resultado.erros.append(
            f"Reconheci {nome_arquivo!r} como uma planilha Excel, mas não como "
            "um dos arquivos da corretora. Esperava a 'PosicaoDetalhada' ou o "
            "'Extrato' da conta XP."
        )
        return resultado, None

    amostra = dados_bytes[:2000].decode("utf-8", errors="ignore")
    tipo = detectar_tipo(nome_arquivo, amostra)

    if tipo == TIPO_OFX:
        return extrato_ofx.ler_bytes(dados_bytes, nome_arquivo), tipo

    texto = dados_bytes.decode("utf-8-sig", errors="replace")

    if tipo == TIPO_FATURA:
        return fatura_csv.ler_texto(texto, nome_arquivo, mes_competencia), tipo
    if tipo == TIPO_EXTRATO_CSV:
        return extrato_csv.ler_texto(texto, nome_arquivo), tipo

    resultado = ResultadoLeitura()
    resultado.erros.append(
        f"Não reconheci o tipo do arquivo {nome_arquivo!r}. "
        "Esperava uma fatura CSV (com coluna Estabelecimento), um extrato CSV "
        "(com colunas Descrição e Saldo), um extrato OFX, ou uma planilha "
        "da corretora (PosicaoDetalhada.xlsx / Extrato ....xlsx)."
    )
    return resultado, None


def assinatura_fatura(mes_competencia, data, descricao, portador, valor,
                      parcela_atual, parcela_total) -> tuple:
    """Os campos que identificam uma linha de FATURA, na ordem que o hash usa.

    ESTA FUNCAO EXISTE PARA SER A DEFINICAO UNICA. Ela e chamada de dois
    lugares que precisam concordar ou a deduplicacao para de funcionar:

        `montar_ids`          quando o arquivo e lido
        migracao 14 (banco)   quando os ids gravados sao recalculados

    Enquanto a regra estava escrita duas vezes, as duas versoes divergiram —
    e o resultado foram ids em colisao, que a migracao 14 teve de reparar.

    POR QUE `parcela_atual` E `parcela_total`, E NAO O `parcela_texto`
    ------------------------------------------------------------------
    Ate 2026-08-25 a assinatura usava o TEXTO CRU da coluna Parcela ("3 de 3",
    "-", " de 1"). Parecia mais fiel ao arquivo, e tinha um defeito silencioso:

        **`parcela_texto` NAO E UMA COLUNA DE `lancamentos`.**

    Ele e lido, entra no hash e e descartado na gravacao. Ou seja, a impressao
    digital de uma linha dependia de um dado que o banco nao guarda — e por
    isso nunca foi reproduzivel a partir do proprio banco. Nenhuma conferencia
    e nenhuma migracao conseguia recalcular o id de uma linha ja gravada.

    `parcela_atual` e `parcela_total` sao o MESMO dado ja normalizado por
    `parse_parcela`, e esses o banco guarda. Nao ha perda de poder de
    distincao: dois textos diferentes que significam a mesma parcela ("-" e
    " de 1" dao os dois (1, 1)) descrevem a mesma coisa, e duas linhas
    realmente distintas continuam separadas pelo contador de ocorrencia.

    O ganho e que a identidade de uma linha passa a depender do dado, e nao da
    forma como o banco escreveu o texto naquele mes.
    """
    return (mes_competencia, data, descricao, portador, valor,
            int(parcela_atual or 1), int(parcela_total or 1))


def montar_ids(linhas: list[dict]) -> list[str]:
    """Gera o id_unico de cada linha lida.

    O contador de ocorrencia (`ocorrencias`) e a parte sutil: ele so entra em
    acao quando duas linhas do MESMO arquivo geram a mesma assinatura. A
    primeira recebe o sufixo 1, a segunda o sufixo 2, e as duas passam a ser
    distintas. Como o contador e reconstruido da mesma forma toda vez que o
    arquivo e lido, reimportar o mesmo arquivo continua gerando os mesmos ids
    — e a deduplicacao segue funcionando.
    """
    ocorrencias: Counter = Counter()
    ids = []

    for linha in linhas:
        if linha.get("fitid"):
            ids.append(chave_hash("OFX", linha["fitid"]))
            continue

        if linha.get("origem") == config.ORIGEM_FATURA:
            assinatura = assinatura_fatura(
                linha["mes_competencia"], linha["data"], linha["descricao"],
                linha.get("portador"), linha["valor"],
                linha.get("parcela_atual", 1), linha.get("parcela_total", 1),
            )
        else:
            assinatura = (
                linha["data"], linha.get("hora"), linha["descricao"], linha["valor"],
            )

        ocorrencias[assinatura] += 1
        ids.append(chave_hash(*assinatura, ocorrencias[assinatura]))

    return ids


def assinatura_cruzada(data: str, valor: float, descricao: str) -> str:
    """Impressao digital que NAO depende de qual arquivo a transacao veio.

    POR QUE ISSO PRECISA EXISTIR — o bug que este projeto quase teve:

    Os seus extratos se sobrepoem. O CSV vai ate 06/08 e o OFX comeca em
    24/07; 21 transacoes, somando R$ ···· estao NOS DOIS ARQUIVOS.

    Mas o id_unico do OFX e feito a partir do FITID, e o do CSV a partir de
    data+hora+descricao+valor. A MESMA transacao gera ids diferentes conforme
    o arquivo de origem. Ou seja: a deduplicacao normal nao veria problema
    nenhum, e importar os dois arquivos contaria R$ ···· duas vezes.

    A solucao e uma segunda impressao digital, feita so com o que as duas
    fontes TEM EM COMUM:

        data + valor + comeco da descricao

    A hora fica de fora porque o OFX nao a fornece. A descricao e cortada em
    40 caracteres porque os dois formatos as vezes truncam o texto em
    tamanhos diferentes ("Pix recebido de Daniel Juchem Bermudez" x
    "Pix recebido de Daniel Juchem Berm").
    """
    from financas.formato import normalizar_texto

    return chave_hash("CRUZ", data, f"{round(float(valor), 2):.2f}",
                      normalizar_texto(descricao)[:40])


def assinaturas_cruzadas_no_banco(data_inicio: str, data_fim: str) -> Counter:
    """Conta quantas vezes cada assinatura cruzada ja existe no banco.

    Devolve um Counter (dicionario de contagens) e nao um conjunto, DE
    PROPOSITO. Se voce mandou dois Pix de R$ ···· para a mesma pessoa no mesmo
    dia, os dois sao transacoes de verdade e os dois devem entrar. Contando
    quantas ja existem, marcamos como duplicata exatamente essa quantidade e
    deixamos passar o excedente.
    """
    linhas = banco.consultar(
        "SELECT data, valor, descricao FROM lancamentos "
        "WHERE origem = ? AND data BETWEEN ? AND ?",
        (config.ORIGEM_EXTRATO, data_inicio, data_fim),
    )
    return Counter(
        assinatura_cruzada(linha["data"], linha["valor"], linha["descricao"])
        for linha in linhas
    )


def ids_existentes(ids: list[str]) -> set[str]:
    """Consulta quais desses ids ja estao no banco.

    Fazemos UMA consulta com todos os ids em vez de uma por linha. Com 500
    linhas isso e a diferenca entre 500 idas ao banco e uma so.

    O SQLite tem limite de variaveis por consulta (999 por padrao), entao
    quebramos em blocos de 500 por seguranca.
    """
    if not ids:
        return set()

    encontrados: set[str] = set()
    for inicio in range(0, len(ids), 500):
        bloco = ids[inicio:inicio + 500]
        marcadores = ",".join("?" * len(bloco))
        linhas = banco.consultar(
            f"SELECT id_unico FROM lancamentos WHERE id_unico IN ({marcadores})",
            bloco,
        )
        encontrados.update(linha["id_unico"] for linha in linhas)
    return encontrados


def preparar(resultado: ResultadoLeitura,
             conjunto_regras: regras.ConjuntoDeRegras | None = None) -> pd.DataFrame:
    """Transforma o resultado da leitura na tabela que voce revisa na tela.

    Devolve um DataFrame com todas as colunas do lancamento MAIS:
        id_unico   a impressao digital
        status     NOVO ou DUPLICADO
        regra      qual regra classificou (vazio quando nenhuma casou)
        importar   True/False — o que a tela usa como caixa de selecao

    Repare que `importar` ja vem marcado so nos NOVOS. Assim o caminho comum
    (importar tudo que e novo) e um clique so, e voce ainda pode desmarcar
    linhas antes de gravar.
    """
    colunas = ["importar", "status", "motivo", "data", "hora", "mes_competencia",
               "descricao", "portador", "valor", "categoria", "tipo",
               "natureza", "origem", "parcela_atual", "parcela_total",
               "parcela_texto", "saldo_apos", "fitid", "regra", "id_unico"]

    if not resultado.linhas:
        return pd.DataFrame(columns=colunas)

    if conjunto_regras is None:
        conjunto_regras = regras.carregar_regras()

    ids = montar_ids(resultado.linhas)
    ja_existem = ids_existentes(ids)

    linhas_extrato = [l for l in resultado.linhas
                      if l.get("origem") == config.ORIGEM_EXTRATO]
    if linhas_extrato:
        datas = [l["data"] for l in linhas_extrato]
        restantes = assinaturas_cruzadas_no_banco(min(datas), max(datas))
    else:
        restantes = Counter()

    registros = []
    for linha, id_unico in zip(resultado.linhas, ids):
        classificacao = regras.classificar(linha, conjunto_regras)

        categoria_do_portador = regras.categoria_por_portador(linha.get("portador"))
        if categoria_do_portador:
            classificacao = regras.Classificacao(
                categoria=categoria_do_portador,
                tipo=classificacao.tipo,
                natureza=classificacao.natureza,
                regra=f"portador: {linha.get('portador')}",
            )

        duplicado = id_unico in ja_existem
        motivo = "id igual" if duplicado else ""

        if not duplicado and linha.get("origem") == config.ORIGEM_EXTRATO:
            cruzada = assinatura_cruzada(
                linha["data"], linha["valor"], linha["descricao"])
            if restantes.get(cruzada, 0) > 0:
                restantes[cruzada] -= 1
                duplicado = True
                motivo = "já existe vinda de outro arquivo"

        registros.append({
            "importar": not duplicado,
            "status": STATUS_DUPLICADO if duplicado else STATUS_NOVO,
            "motivo": motivo,
            "data": linha["data"],
            "hora": linha.get("hora"),
            "mes_competencia": linha["mes_competencia"],
            "descricao": linha["descricao"],
            "portador": linha.get("portador"),
            "valor": linha["valor"],
            "categoria": classificacao.categoria,
            "tipo": classificacao.tipo,
            "natureza": classificacao.natureza,
            "origem": linha["origem"],
            "parcela_atual": linha.get("parcela_atual", 1),
            "parcela_total": linha.get("parcela_total", 1),
            "parcela_texto": linha.get("parcela_texto"),
            "saldo_apos": linha.get("saldo_apos"),
            "fitid": linha.get("fitid"),
            "regra": classificacao.regra or "",
            "id_unico": id_unico,
        })

    return pd.DataFrame(registros, columns=colunas)


def resumo_previa(df: pd.DataFrame) -> dict:
    """Os numeros que aparecem acima da tabela de previa."""
    if df.empty:
        return {"total": 0, "novos": 0, "duplicados": 0, "sem_regra": 0,
                "valor_novos": 0.0, "entradas": 0.0, "saidas": 0.0}

    novos = df[df["status"] == STATUS_NOVO]
    return {
        "total": int(len(df)),
        "novos": int(len(novos)),
        "duplicados": int((df["status"] == STATUS_DUPLICADO).sum()),
        "sem_regra": int((df["regra"] == "").sum()),
        "valor_novos": float(novos["valor"].sum()) if not novos.empty else 0.0,
        "entradas": float(novos[novos["valor"] > 0]["valor"].sum()) if not novos.empty else 0.0,
        "saidas": float(novos[novos["valor"] < 0]["valor"].sum()) if not novos.empty else 0.0,
    }


def _chave_parcelamento(descricao: str, mes: str, atual: int, total: int) -> str | None:
    """Monta a chave que junta as parcelas da mesma compra.

    Repete a mesma regra usada na migracao (migracao/carregar.py) para que
    lancamento migrado e lancamento importado caiam na MESMA chave — senao o
    mesmo parcelamento apareceria partido em dois.
    """
    from financas.formato import mes_para_indice, normalizar_texto

    if total <= 1:
        return None
    indice = mes_para_indice(mes)
    if indice is None:
        return None
    return f"{normalizar_texto(descricao)}|{total}|{indice - (atual - 1)}"


def gravar(df: pd.DataFrame, nome_arquivo: str = "",
           sha256: str = "", tipo: str = "") -> dict:
    """Grava no banco as linhas marcadas para importar.

    Devolve {gravados, ignorados, duplicados_no_banco}.

    `gravados` vem da contagem REAL de linhas na tabela antes e depois, e nao
    do tamanho da lista enviada. Assim, se o banco recusar alguma duplicata
    que passou pela nossa checagem, o numero mostrado continua sendo a
    verdade.
    """
    if df.empty:
        return {"gravados": 0, "ignorados": 0, "duplicados_no_banco": 0}

    selecionadas = df[df["importar"].fillna(False).astype(bool)]
    if selecionadas.empty:
        return {"gravados": 0, "ignorados": int(len(df)), "duplicados_no_banco": 0}

    id_cartao = banco.id_conta("Cartão XP")
    id_conta_corrente = banco.id_conta("Conta Corrente XP")
    carimbo = banco.agora()

    registros = []
    for _, linha in selecionadas.iterrows():
        origem = linha["origem"]
        conta_id = id_cartao if origem == config.ORIGEM_FATURA else id_conta_corrente
        atual = int(linha.get("parcela_atual") or 1)
        total = int(linha.get("parcela_total") or 1)

        registros.append((
            linha["id_unico"],
            linha["data"],
            linha.get("hora"),
            linha["mes_competencia"],
            linha["descricao"],
            linha.get("portador"),
            float(linha["valor"]),
            linha.get("categoria"),
            linha.get("tipo"),
            linha.get("natureza"),
            origem,
            conta_id,
            atual,
            total,
            _chave_parcelamento(linha["descricao"], linha["mes_competencia"], atual, total),
            linha.get("fitid"),
            linha.get("saldo_apos") if pd.notna(linha.get("saldo_apos")) else None,
            None,
            linha.get("regra") or None,
            carimbo,
            carimbo,
        ))

    antes = banco.contar("lancamentos")
    banco.executar_muitos(
        """INSERT OR IGNORE INTO lancamentos
           (id_unico, data, hora, mes_competencia, descricao, portador, valor,
            categoria, tipo, natureza, origem, conta_id,
            parcela_atual, parcela_total, chave_parcelamento,
            fitid, saldo_apos, observacao, regra_aplicada, criado_em, atualizado_em)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        registros,
    )
    depois = banco.contar("lancamentos")
    gravados = depois - antes

    if nome_arquivo:
        banco.executar(
            """INSERT INTO arquivos_importados
               (nome, sha256, tipo, mes_referencia, linhas_lidas, linhas_novas,
                linhas_dup, importado_em)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                nome_arquivo, sha256, tipo,
                str(selecionadas["mes_competencia"].iloc[0]),
                int(len(df)), gravados,
                int((df["status"] == STATUS_DUPLICADO).sum()),
                carimbo,
            ),
        )

    return {
        "gravados": gravados,
        "ignorados": int(len(df) - len(selecionadas)),
        "duplicados_no_banco": int(len(selecionadas) - gravados),
    }


def ja_importado_por_hash(digest: str) -> dict | None:
    """Diz se um arquivo com ESTE conteudo ja foi importado antes.

    Recebe o hash pronto, e nao um caminho, porque o upload do Streamlit
    entrega bytes na memoria — nunca um arquivo em disco. Enquanto so existia
    a versao por caminho, o aviso de "voce ja importou isto" nunca aparecia
    para quem sobe pelo uploader, que e o caminho normal de uso.

    A guarda `sha256 <> ''` nao e decoracao: ate 2026-08-25 o ramo do upload
    gravava o hash em branco, e sem ela um digest vazio casaria com todos
    esses registros de uma vez.
    """
    if not digest:
        return None
    linha = banco.consultar_um(
        "SELECT * FROM arquivos_importados WHERE sha256 = ? AND sha256 <> '' "
        "ORDER BY id DESC LIMIT 1",
        (digest,),
    )
    return dict(linha) if linha else None


def historico(limite: int = 50) -> pd.DataFrame:
    """As ultimas importacoes feitas."""
    return banco.df(
        "SELECT nome, tipo, mes_referencia, linhas_lidas, linhas_novas, "
        "linhas_dup, importado_em FROM arquivos_importados "
        "ORDER BY id DESC LIMIT ?",
        (limite,),
    )


EXTENSOES_ACEITAS = (".csv", ".ofx", ".xlsx")


def arquivos_da_pasta(pasta) -> list[Path]:
    """Lista os arquivos importaveis de uma pasta, do mais novo para o mais velho."""
    pasta = Path(pasta)
    if not pasta.is_dir():
        return []
    encontrados = [
        caminho for caminho in pasta.iterdir()
        if caminho.is_file() and caminho.suffix.lower() in EXTENSOES_ACEITAS
    ]
    return sorted(encontrados, key=lambda c: c.stat().st_mtime, reverse=True)


def estado_da_pasta(pasta) -> pd.DataFrame:
    """Cada arquivo da pasta com a marca de "ja entrou" ou "ainda nao".

    Colunas: nome, caminho, tamanho, modificado, importado, importado_em,
             linhas_novas, tipo

    POR QUE ISTO PRECISOU EXISTIR
    -----------------------------
    A tela listava os arquivos da pasta e so dizia "voce ja importou isto"
    DEPOIS de voce escolher um e ele ser lido. Com 71 arquivos guardados, achar
    o que ainda falta virava tentativa e erro — abrir, ler o aviso, voltar.

    A marca sai do mesmo hash que a deduplicacao usa, entao ela nao e um
    palpite pelo nome: um arquivo renomeado continua sendo reconhecido, e dois
    downloads do mesmo extrato com nomes diferentes aparecem os dois como
    importados.

    Ler 71 arquivos para calcular hash custa tempo, e por isso quem chama deve
    guardar o resultado em cache — `ui/estado.arquivos_da_pasta` faz isso,
    com o tamanho e a data de modificacao na chave.
    """
    colunas = ["nome", "caminho", "tamanho", "modificado", "importado",
               "importado_em", "linhas_novas", "tipo"]
    encontrados = arquivos_da_pasta(pasta)
    if not encontrados:
        return pd.DataFrame(columns=colunas)

    linhas = []
    for caminho in encontrados:
        informacao = caminho.stat()
        try:
            digest = hash_arquivo(caminho)
        except OSError:
            digest = ""
        registro = ja_importado_por_hash(digest) if digest else None
        linhas.append({
            "nome": caminho.name,
            "caminho": str(caminho),
            "tamanho": int(informacao.st_size),
            "modificado": datetime.fromtimestamp(
                informacao.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "importado": registro is not None,
            "importado_em": registro["importado_em"] if registro else None,
            "linhas_novas": registro["linhas_novas"] if registro else None,
            "tipo": registro["tipo"] if registro else None,
        })
    return pd.DataFrame(linhas, columns=colunas)


def cobertura_por_tipo() -> pd.DataFrame:
    """O que ja entrou de cada tipo de arquivo, e quando foi a ultima vez.

    Colunas: tipo, arquivos, linhas_novas, ultimo_arquivo, ultima_importacao

    Responde "o que falta este mes?" sem inventar um calendario: o app nao sabe
    quando a sua fatura fecha nem quando voce baixa a posicao. O que ele sabe e
    quando cada tipo entrou pela ultima vez — e um tipo que nao aparece ha dois
    meses salta aos olhos sozinho.
    """
    return banco.df(
        """SELECT tipo,
                  COUNT(*)                AS arquivos,
                  SUM(linhas_novas)       AS linhas_novas,
                  MAX(nome)               AS ultimo_arquivo,
                  MAX(importado_em)       AS ultima_importacao
             FROM arquivos_importados
            GROUP BY tipo
            ORDER BY ultima_importacao DESC""")


def detectar_tipo_excel(dados_bytes: bytes, nome_arquivo: str = "") -> str | None:
    """Diz se um .xlsx e a POSICAO da carteira ou o EXTRATO da corretora.

    Olha o conteudo, nao so o nome. Um xlsx e um arquivo ZIP: os primeiros
    bytes sao sempre "PK", entao nao da para procurar palavra-chave no texto
    como fazemos com CSV e OFX. Abrimos a planilha e perguntamos a ela:

        aba chamada "Sua carteira"        -> posicao
        celula "Movimentação" + "Valor (R$)" -> extrato

    Se a planilha nao abrir, cai para o nome do arquivo, que acerta na maioria
    das vezes ("PosicaoDetalhada.xlsx", "Extrato 12345678 ....xlsx").
    """
    import io

    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(dados_bytes), data_only=True, read_only=True)
        try:
            if "Sua carteira" in wb.sheetnames:
                return TIPO_POSICAO_XP
            ws = wb.worksheets[0]
            for indice, linha in enumerate(ws.iter_rows(values_only=True)):
                if indice > 30:
                    break
                celulas = {str(c).strip() for c in linha if c is not None}
                if "Movimentação" in celulas and "Valor (R$)" in celulas:
                    return TIPO_EXTRATO_XP
                if any(str(c).strip().startswith("Total investido")
                       for c in celulas):
                    return TIPO_POSICAO_XP
        finally:
            wb.close()
    except Exception:
        pass

    nome = (nome_arquivo or "").lower()
    if "posicao" in nome or "posição" in nome:
        return TIPO_POSICAO_XP
    if "extrato" in nome:
        return TIPO_EXTRATO_XP
    return None


def gravar_arquivo_xp(resultado: ResultadoLeitura, tipo: str,
                      nome_arquivo: str = "", sha256: str = "") -> dict:
    """Grava um arquivo da corretora e registra no historico de importacao.

    Devolve um resumo pronto para a tela, com as mesmas chaves nos dois casos
    (`gravados`, `ignorados`) mais o que for especifico de cada tipo.
    """
    from financas.calculos import investimentos as calculo_investimentos

    if tipo == TIPO_POSICAO_XP:
        resumo = calculo_investimentos.gravar_posicao(resultado)
        gravados = resumo["criados"] + resumo["atualizados"]
        ignorados = 0
        mes_referencia = resumo["mes"]
    elif tipo == TIPO_EXTRATO_XP:
        resumo = calculo_investimentos.gravar_movimentos(resultado)
        gravados = resumo["gravados"]
        ignorados = resumo["ignorados"]
        mes_referencia = (resultado.meta.get("meses") or [None])[-1]
    else:
        return {"gravados": 0, "ignorados": 0}

    if nome_arquivo:
        banco.executar(
            """INSERT INTO arquivos_importados
               (nome, sha256, tipo, mes_referencia, linhas_lidas, linhas_novas,
                linhas_dup, importado_em)
               VALUES (?,?,?,?,?,?,?,?)""",
            (nome_arquivo, sha256, tipo, mes_referencia,
             len(resultado.linhas), gravados, ignorados, banco.agora()),
        )

    resumo["gravados"] = gravados
    resumo["ignorados"] = ignorados
    return resumo
