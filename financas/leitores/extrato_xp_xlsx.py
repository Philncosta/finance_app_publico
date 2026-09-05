"""
extrato_xp_xlsx.py — Le o extrato da CONTA DE INVESTIMENTO (XP), em Excel.
==============================================================================

NAO CONFUNDA COM O EXTRATO DO BANCO
-----------------------------------
Sao duas contas diferentes, e este leitor cuida da segunda:

  * `extrato_csv.py` / `extrato_ofx.py` leem a sua CONTA CORRENTE. O que sai
    de la vira despesa no Dashboard.
  * este arquivo le a CONTA DA CORRETORA. O que acontece aqui — comprar um
    titulo, receber juros, pagar IRRF — nao e receita nem despesa do seu mes;
    e dinheiro trocando de lugar DENTRO do seu patrimonio.

Por isso estes lancamentos vao para a tabela `investimentos_movimentos`, e
nunca para `lancamentos`. Se fossem para `lancamentos`, cada compra de titulo
viraria despesa e o Dashboard diria que voce gastou R$ ····.

O ARQUIVO
---------
    r2..r12   cabecalho (titulo, periodo, conta, saldo projetado)
    r14       Movimentacao | Liquidacao | Lancamento |  | Valor (R$) | Saldo (R$)
    r15+      2026-08-20 | 2026-08-20 | COMPRA TESOURO DIRETO... | | -477.57 | 22.74

As duas datas tem significados diferentes: **Movimentacao** e quando a ordem
aconteceu, **Liquidacao** e quando o dinheiro efetivamente entrou ou saiu.
Para efeito de mes usamos a movimentacao, que e a que casa com o extrato da
conta corrente; a liquidacao fica guardada para consulta.

DE ONDE SAI O `tipo_movimento`
------------------------------
O extrato nao tem coluna de tipo — so o texto do lancamento. A tabela `REGRAS`
traduz esse texto. A ORDEM das regras importa mais que o conteudo, porque
varios textos se sobrepoem:

    "IRRF S/RESGATE FUNDOS - Trend DI"       contem "RESGATE", mas e imposto
    "TED APLICACAO FUNDOS Trend Investback"  contem "INVESTBACK", mas e compra

Se "RESGATE" fosse testado antes de "IRRF", o imposto viraria venda. Por isso
as regras mais especificas vem primeiro, e a primeira que casar decide.
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path

from financas.formato import (chave_hash, mes_de, normalizar_texto, parse_brl,
                              parse_data)
from financas.leitores.base import ResultadoLeitura

warnings.filterwarnings("ignore", message=".*extension is not supported.*")

COL_DATA = "Movimentação"
COL_LIQUIDACAO = "Liquidação"
COL_DESCRICAO = "Lançamento"
COL_VALOR = "Valor (R$)"
COL_SALDO = "Saldo (R$)"

MARCA_FUTUROS = "Lançamentos futuros"

REGRAS = [
    ("IRRF",                          "imposto"),
    ("IOF",                           "imposto"),
    ("OPERACOES EM BOLSA",            "taxa"),
    ("RECEBIDA DA CONTA DIGITAL",     "aporte"),
    ("ENVIADA PARA A CONTA DIGITAL",  "resgate"),
    ("RETIRADA EM C/C",               "resgate"),
    ("COMPRA",                        "compra"),
    ("APLICACAO FUNDOS",              "compra"),
    ("RESGATE",                       "venda"),
    ("REPASSE DE JUROS",              "juros"),
    ("DIVIDENDOS",                    "dividendo"),
    ("JUROS SOBRE CAPITAL",           "dividendo"),
    ("INVESTBACK",                    "rendimento"),
]

DESCRICAO_TIPOS = {
    "aporte":     "dinheiro que voce mandou da conta corrente para a corretora",
    "resgate":    "dinheiro que voltou da corretora para a conta corrente",
    "compra":     "compra de titulo ou aplicacao em fundo",
    "venda":      "resgate de fundo ou venda de titulo (vira saldo em conta)",
    "juros":      "cupom de juros pago por um titulo",
    "dividendo":  "dividendo ou juros sobre capital proprio de acao",
    "rendimento": "rendimento creditado (cashback do Investback)",
    "imposto":    "IRRF ou IOF retido na fonte",
    "taxa":       "taxa de corretagem ou de bolsa",
    "outro":      "nao reconhecido — confira na tela de Movimentacoes",
}

TIPOS_EXTERNOS = ("aporte", "resgate")


def classificar_movimento(descricao: str) -> str:
    """Traduz o texto do lancamento num tipo. Ver `REGRAS` e o topo do arquivo."""
    texto = normalizar_texto(descricao)
    for trecho, tipo in REGRAS:
        if trecho in texto:
            return tipo
    return "outro"


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _achar_cabecalho(linhas) -> int | None:
    """Devolve o indice da linha de cabecalho, ou None se nao houver.

    Procuramos em vez de fixar "linha 14" porque a corretora pode mudar o
    tamanho do cabecalho sem aviso — e ai o leitor pararia de funcionar sem
    ninguem entender o motivo.
    """
    for indice, linha in enumerate(linhas):
        celulas = [_texto(c) for c in linha]
        if COL_DATA in celulas and COL_VALOR in celulas:
            return indice
    return None


def ler_planilha(ws) -> ResultadoLeitura:
    """Le a aba do extrato ja aberta e devolve o resultado normalizado."""
    resultado = ResultadoLeitura()
    linhas = [list(linha) for linha in ws.iter_rows(values_only=True)]

    indice_cabecalho = _achar_cabecalho(linhas)
    if indice_cabecalho is None:
        resultado.erros.append(
            "Não achei a linha de cabeçalho (Movimentação / Valor (R$)). "
            "Confira se é mesmo o extrato da conta XP em Excel."
        )
        return resultado

    colunas = {nome: i for i, nome in
               enumerate(_texto(c) for c in linhas[indice_cabecalho]) if nome}

    topo = " ".join(_texto(c) for linha in linhas[:indice_cabecalho] for c in linha)
    conta = re.search(r"Conta XP:\s*(\d+)", topo)
    periodo = re.search(r"De:\s*(\d{2}/\d{2}/\d{4})\s*At[ée]:\s*(\d{2}/\d{2}/\d{4})", topo)

    saldo_projetado = None
    for linha in linhas[:indice_cabecalho]:
        celulas = [_texto(c) for c in linha]
        if any(c.startswith("Saldo total projetado") for c in celulas):
            for celula_texto in reversed(celulas):
                if parse_brl(celula_texto) is not None:
                    saldo_projetado = parse_brl(celula_texto)
                    break
            break

    def campo(linha, nome):
        """Le uma coluna da linha pelo NOME, nao pela posicao."""
        posicao = colunas.get(nome)
        if posicao is None or posicao >= len(linha):
            return None
        return linha[posicao]

    fim_realizados = len(linhas)
    for indice in range(indice_cabecalho + 1, len(linhas)):
        if any(_texto(c) == MARCA_FUTUROS for c in linhas[indice]):
            fim_realizados = indice
            break

    ignoradas = 0
    for numero, linha in enumerate(linhas[indice_cabecalho + 1:fim_realizados],
                                   start=indice_cabecalho + 2):
        if not any(_texto(c) for c in linha):
            continue

        data = parse_data(campo(linha, COL_DATA))
        valor = parse_brl(campo(linha, COL_VALOR))
        descricao = _texto(campo(linha, COL_DESCRICAO))

        if data is None or valor is None or not descricao:
            ignoradas += 1
            continue

        liquidacao = parse_data(campo(linha, COL_LIQUIDACAO))
        saldo = parse_brl(campo(linha, COL_SALDO))

        id_unico = chave_hash(data.isoformat(), descricao, round(valor, 2),
                              round(saldo, 2) if saldo is not None else "")

        resultado.linhas.append({
            "id_unico": id_unico,
            "data": data.isoformat(),
            "liquidacao": liquidacao.isoformat() if liquidacao else None,
            "mes_competencia": mes_de(data),
            "descricao": descricao,
            "valor": valor,
            "saldo_apos": saldo,
            "tipo_movimento": classificar_movimento(descricao),
            "linha_arquivo": numero,
        })

    if not resultado.linhas:
        resultado.erros.append("O extrato não tem nenhuma movimentação.")
        return resultado

    if ignoradas:
        resultado.avisos.append(
            f"{ignoradas} linha(s) ignorada(s) por não terem data ou valor."
        )

    futuros = 0
    for linha in linhas[fim_realizados:]:
        if (parse_data(campo(linha, COL_DATA)) is not None
                and parse_brl(campo(linha, COL_VALOR)) is not None
                and _texto(campo(linha, COL_DESCRICAO))):
            futuros += 1
    if futuros:
        resultado.avisos.append(
            f"{futuros} lançamento(s) futuro(s) agendado(s) não foram "
            f"importados — eles entram quando de fato acontecerem."
        )

    nao_reconhecidos = [l["descricao"] for l in resultado.linhas
                        if l["tipo_movimento"] == "outro"]
    if nao_reconhecidos:
        exemplos = "; ".join(sorted(set(nao_reconhecidos))[:3])
        resultado.avisos.append(
            f"{len(nao_reconhecidos)} movimentação(ões) sem tipo reconhecido "
            f"(ex.: {exemplos}). Entram como 'outro' e podem ser "
            f"reclassificadas na tela de Movimentações."
        )

    datas = [l["data"] for l in resultado.linhas]
    por_tipo: dict[str, dict] = {}
    for l in resultado.linhas:
        alvo = por_tipo.setdefault(l["tipo_movimento"], {"quantidade": 0, "soma": 0.0})
        alvo["quantidade"] += 1
        alvo["soma"] += l["valor"]

    resultado.meta = {
        "tipo": "Extrato da conta de investimento",
        "conta": conta.group(1) if conta else None,
        "inicio": min(datas),
        "fim": max(datas),
        "periodo": (f"{periodo.group(1)} a {periodo.group(2)}" if periodo
                    else f"{min(datas)} a {max(datas)}"),
        "saldo_projetado": saldo_projetado,
        "meses": sorted({l["mes_competencia"] for l in resultado.linhas}),
        "por_tipo": por_tipo,
    }
    return resultado


def ler_bytes(dados: bytes, nome_arquivo: str = "") -> ResultadoLeitura:
    """Le o extrato a partir dos bytes (o que o upload do Streamlit entrega)."""
    import io

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(dados), data_only=True, read_only=True)
    except Exception as erro:
        resultado = ResultadoLeitura()
        resultado.erros.append(f"Não consegui abrir a planilha: {erro}")
        return resultado

    try:
        return ler_planilha(wb.worksheets[0])
    finally:
        wb.close()


def ler_arquivo(caminho) -> ResultadoLeitura:
    """Le o extrato direto de um arquivo em disco."""
    caminho = Path(caminho)
    try:
        return ler_bytes(caminho.read_bytes(), caminho.name)
    except OSError as erro:
        resultado = ResultadoLeitura()
        resultado.erros.append(f"Não consegui abrir o arquivo: {erro}")
        return resultado
