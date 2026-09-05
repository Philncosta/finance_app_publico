"""
posicao_xp.py — Le a posicao da carteira exportada pela corretora (XP).
==============================================================================

O ARQUIVO DE ENTRADA
--------------------
`PosicaoDetalhada.xlsx` (posicao de hoje) ou
`PosicaoDetalhadaHistorica_DD_MM_AAAA.xlsx` (posicao de uma data passada).

Uma aba so, chamada "Sua carteira", organizada em BLOCOS:

    r1   Conta: 12345678 | 22/08/2026, 07:27
    r3   ...  | Total investido | Saldo Disponível | ...
    r4   R$ ···· | R$ ···· | R$ ···· | ...

    r6   Fundos de Investimentos                            R$ ····   <- grupo
    r8   26,2% | Pós-Fixado | Posição | % Alocação | ...                    <- colunas
    r9   Trend DI FIC RF Simples RL | R$ ···· | 25,96% | ...           <- ativo

    r13  Tesouro Direto                                    R$ ····   <- grupo
    r15  73,8% | Pós-Fixado | Posição | % Alocação | Total aplicado | ...   <- colunas
    r16  NTN-B ago/2060 | R$ ···· | 28,83% | ... | 15/08/2060          <- ativo

AS DUAS ARMADILHAS DESTE ARQUIVO
--------------------------------
**1. As colunas MUDAM de bloco para bloco.**

    Fundos:   Posição | % Alocação | Rent. Líquida | Rent. Bruta | Valor aplicado | Valor líquido
    Tesouro:  Posição | % Alocação | Total aplicado | Qtd. | Disponível | Vencimento

Ler com posicao fixa daria "Rentabilidade Líquida = 15/08/2060" no bloco do
Tesouro. Por isso o leitor monta um mapa de colunas A CADA subcabecalho, e usa
esse mapa so ate o proximo bloco.

**2. O rotulo do bloco NAO diz o indexador de verdade.**

No arquivo atual, o bloco do Tesouro esta rotulado "Pós-Fixado" e contem
NTN-B (que e indexada a INFLACAO) junto com LFT (essa sim pos-fixada). No
arquivo historico, os mesmos papeis aparecem em dois blocos separados e
rotulados corretamente.

Ou seja: o rotulo depende de como a corretora agrupou naquele dia. Por isso
guardamos o rotulo apenas como informacao, e a CLASSE de cada papel e deduzida
do NOME dele (`calculos/investimentos.classificar_papel`), que e confiavel.

A DATA DA POSICAO
-----------------
E o dado mais importante do cabecalho, porque decide em que mes o saldo entra.
Os dois formatos:

    "Conta: 12345678 | 22/08/2026, 07:27"                      -> 22/08/2026
    "Conta: ... | Data da consulta: 22/08/2026 |
                  Data da Posição Histórica: 31/07/2026"        -> 31/07/2026

Quando existe "Posição Histórica", e ELA que vale — a data de consulta e so
quando o arquivo foi baixado.
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path

from financas.formato import mes_de, parse_brl, parse_data, vazio
from financas.leitores.base import ResultadoLeitura

warnings.filterwarnings("ignore", message=".*extension is not supported.*")

MARCA_SUBCABECALHO = "Posição"


def _texto(valor) -> str:
    """Converte celula em texto limpo."""
    return "" if valor is None else str(valor).strip()


def _percentual(texto) -> float | None:
    """Converte "25,96%" em 0.2596.

    Guardamos porcentagem como FRACAO em todo o projeto (ver formato.fmt_pct),
    entao dividimos por 100 aqui, na entrada.
    """
    if vazio(texto):
        return None
    numero = parse_brl(str(texto).replace("%", ""))
    return None if numero is None else numero / 100


def _data_da_posicao(cabecalho: str):
    """Extrai a data que vale, tratando os dois formatos do cabecalho."""
    if not cabecalho:
        return None
    historica = re.search(r"Posi[çc][ãa]o Hist[óo]rica:\s*(\d{2}/\d{2}/\d{4})",
                          cabecalho, re.IGNORECASE)
    if historica:
        return parse_data(historica.group(1))
    qualquer = re.search(r"(\d{2}/\d{2}/\d{4})", cabecalho)
    return parse_data(qualquer.group(1)) if qualquer else None


def ler_planilha(ws) -> ResultadoLeitura:
    """Le a aba "Sua carteira" ja aberta e devolve o resultado normalizado.

    Separado de `ler_arquivo` para poder ser testado com uma planilha montada
    na memoria, sem precisar de arquivo em disco.
    """
    resultado = ResultadoLeitura()
    linhas = [
        [cel for cel in linha]
        for linha in ws.iter_rows(values_only=True)
    ]
    if not linhas:
        resultado.erros.append("A planilha está vazia.")
        return resultado

    texto_topo = " ".join(_texto(c) for linha in linhas[:3] for c in linha)
    conta = re.search(r"Conta:\s*(\d+)", texto_topo)
    data_posicao = _data_da_posicao(texto_topo)

    patrimonio = total_investido = saldo_disponivel = None
    for indice, linha in enumerate(linhas[:8]):
        rotulos = [_texto(c) for c in linha]
        if any(r.startswith("Total investido") for r in rotulos) and indice + 1 < len(linhas):
            valores = linhas[indice + 1]
            for posicao, rotulo in enumerate(rotulos):
                bruto = _texto(valores[posicao]) if posicao < len(valores) else ""
                if rotulo.startswith("Total investido"):
                    total_investido = parse_brl(bruto)
                elif rotulo.startswith("Saldo Disponível"):
                    saldo_disponivel = parse_brl(bruto)
                elif "patrim" in rotulo.lower():
                    patrimonio = parse_brl(bruto)
            break

    if data_posicao is None:
        resultado.erros.append(
            "Não achei a data da posição no cabeçalho do arquivo. "
            "Esperava algo como 'Conta: 12345678 | 22/08/2026'."
        )
        return resultado

    grupo = None
    colunas: dict[str, int] = {}
    rotulo_bloco = ""

    for numero, linha in enumerate(linhas, start=1):
        celulas = [_texto(c) for c in linha]
        if not any(celulas):
            continue

        primeira = celulas[0]

        if len(celulas) > 1 and celulas[1] == MARCA_SUBCABECALHO:
            colunas = {nome: i for i, nome in enumerate(celulas) if nome}
            rotulo_bloco = primeira
            continue

        if (primeira and parse_brl(celulas[1] if len(celulas) > 1 else "") is None
                and any("R$" in c for c in celulas[1:])):
            grupo = primeira
            continue

        if not colunas:
            continue

        indice_posicao = colunas[MARCA_SUBCABECALHO]
        bruto = celulas[indice_posicao] if indice_posicao < len(celulas) else ""
        valor = parse_brl(bruto)
        if not primeira or valor is None:
            continue

        def campo(nome_coluna: str):
            """Le uma coluna deste bloco pelo nome, ou None se ele nao existir."""
            posicao = colunas.get(nome_coluna)
            if posicao is None or posicao >= len(celulas):
                return None
            return celulas[posicao] or None

        vencimento = parse_data(campo("Vencimento"))

        aplicado = parse_brl(campo("Total aplicado") or campo("Valor aplicado") or "")
        if aplicado is not None and (aplicado == 0 or abs(aplicado - valor) < 0.01):
            aplicado = None

        resultado.linhas.append({
            "nome": primeira,
            "grupo": grupo or "(sem grupo)",
            "rotulo_bloco": rotulo_bloco,
            "valor": valor,
            "valor_aplicado": aplicado,
            "percentual_corretora": _percentual(campo("% Alocação")),
            "quantidade": parse_brl(campo("Qtd.") or ""),
            "vencimento": vencimento.isoformat() if vencimento else None,
            "rentabilidade_liquida": _percentual(campo("Rentabilidade Líquida")),
            "data_posicao": data_posicao.isoformat(),
            "mes_competencia": mes_de(data_posicao),
            "linha_arquivo": numero,
        })

    if not resultado.linhas:
        resultado.erros.append(
            "Não encontrei nenhum ativo na planilha. Confira se é mesmo o "
            "arquivo 'PosicaoDetalhada' exportado pela corretora."
        )
        return resultado

    soma = sum(l["valor"] for l in resultado.linhas)

    if total_investido is not None and abs(soma - total_investido) > 1:
        resultado.avisos.append(
            f"A soma dos ativos ({soma:,.2f}) não bate com o total investido "
            f"declarado no arquivo ({total_investido:,.2f}). Pode haver um "
            f"tipo de investimento que o leitor ainda não reconhece."
        )

    resultado.meta = {
        "tipo": "Posição da carteira",
        "conta": conta.group(1) if conta else None,
        "data_posicao": data_posicao.isoformat(),
        "mes_competencia": mes_de(data_posicao),
        "periodo": f"posição em {data_posicao.strftime('%d/%m/%Y')}",
        "patrimonio": patrimonio,
        "total_investido": total_investido,
        "saldo_disponivel": saldo_disponivel,
        "soma_ativos": soma,
        "grupos": sorted({l["grupo"] for l in resultado.linhas}),
    }
    return resultado


def ler_bytes(dados: bytes, nome_arquivo: str = "") -> ResultadoLeitura:
    """Le a posicao a partir dos bytes (o que o upload do Streamlit entrega)."""
    import io

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(dados), data_only=True, read_only=True)
    except Exception as erro:
        resultado = ResultadoLeitura()
        resultado.erros.append(f"Não consegui abrir a planilha: {erro}")
        return resultado

    try:
        ws = wb["Sua carteira"] if "Sua carteira" in wb.sheetnames else wb.worksheets[0]
        return ler_planilha(ws)
    finally:
        wb.close()


def ler_arquivo(caminho) -> ResultadoLeitura:
    """Le a posicao direto de um arquivo em disco."""
    caminho = Path(caminho)
    try:
        return ler_bytes(caminho.read_bytes(), caminho.name)
    except OSError as erro:
        resultado = ResultadoLeitura()
        resultado.erros.append(f"Não consegui abrir o arquivo: {erro}")
        return resultado
